import openpyxl
import base64
import hashlib
import json
import requests
import datetime
import random
import urllib
import xlwt
import time
import os
# 计算时间
start = time.time()
tic = lambda: 'at %1.1f seconds' % (time.time() - start)

#全记录
all_excel = xlwt.Workbook(encoding = 'ascii')
worksheet_all = all_excel.add_sheet('sheet1')
#日志错误错误
error_excel = xlwt.Workbook(encoding = 'ascii')
worksheet_error = error_excel.add_sheet('sheet1')

#全记录(包括卡的)
def all_excel_log(iccid,custName,identityCode,log):
   new_row = row-1
   reponse_json = json.loads(log)
   worksheet_all.write(new_row, 0, label = iccid)
   worksheet_all.write(new_row, 1, label = custName)
   worksheet_all.write(new_row, 2, label = identityCode)
   worksheet_all.write(new_row, 3, label = reponse_json['data']['msg'])
   worksheet_all.write(new_row, 4, label = log)
   all_excel.save('my\\success.xls')
#错误的认证记录，方便第二次记录
def write_error_excel_log(iccid,custName,identityCode,file1,file2,file3):
   new_row = row-1
   worksheet_error.write(new_row, 0, label = iccid)
   worksheet_error.write(new_row, 1, label = custName)
   worksheet_error.write(new_row, 2, label = identityCode)
   worksheet_error.write(new_row, 3, label = file1)
   worksheet_error.write(new_row, 4, label = file2)
   worksheet_error.write(new_row, 5, label = file3)
   error_excel.save('my\\3.xlsx')   
def write_error_log(log):
    with open("my\\log\\error.txt",'a+', encoding='utf-8') as f:
      f.write(log+ "\n")
      f.close()
def write_success_log(log):
    with open("my\\log\\success.txt",'a+', encoding='utf-8') as f:
      f.write(log + "\n")
      f.close()
def write_upload_log(iccid,log):
    with open("my\\log\\supload.txt",'a+', encoding='utf-8') as f:
      f.write(log+ "\n")
      f.close()
def get_basefile(imgfile):
    img_file = ".\\images\\"+imgfile
    if os.path.isfile(img_file):
      with open(img_file, 'rb') as f:
         base64_data = base64.b64encode(f.read())
         base64file = base64_data.decode()
      return base64file
    else:
      return ''
def get_base64file(basefile):
    md5 = hashlib.md5()
    md5.update(basefile.encode("utf-8"))
    return md5.hexdigest()
def get_imgid(iccid,imgfile,basefile,md5basefile):
    file_type = imgfile[-4:]
    dt = datetime.datetime.now()
    filename = dt.strftime("%Y%m%d%H%M%S")+random.choice('abcdefghimnz')+random.choice('abcdefghimnz')+file_type
    values = {'filename':filename,'basefile':basefile,'md5basefile':md5basefile}
    values_json = json.dumps(values)
    url = "http://59.36.143.49:8080/smapi/getPhotoCode"
    req = requests.post(url, data=values_json)
    change = req.json()
    new_req = json.dumps(change, ensure_ascii=False)
    write_upload_log(iccid,new_req)
    return new_req
def get_status(iccid,identitycode,custname,filecategorystr):
    custname = urllib.parse.quote(custname)
    param = "?iccid="+iccid+"&identitycode="+identitycode+"&custname="+custname+"&filecategorystr="+filecategorystr
    url = "http://59.36.143.49:8080/smapi/getIdentityAuth"+param
    req = requests.get(url)
    change = req.json()
    new_req = json.dumps(change, ensure_ascii=False)
    return new_req
print('Started: ', tic())
# 打开excel文件,获取工作簿对象
wb = openpyxl.load_workbook('3.xlsx')
ws = wb.active  # 当前活跃的表单

filecategorystr = ''
first_name =  ''
first_idcode = ''
for row in range(1,ws.max_row+1):
   iccid = ws.cell(row = row, column = 1).value #通过行列读
   if  iccid:
       custname = ws.cell(row = row, column = 2).value 
       idcode = ws.cell(row = row, column = 3).value
       file1 = ws.cell(row = row, column = 4).value 
       file2 = ws.cell(row = row, column = 5).value
       file3 = ws.cell(row = row, column = 6).value 
       #存在相同的
       if first_name == custname and first_idcode == idcode:
          if  filecategorystr !='':
            status_json = get_status(iccid,idcode,custname,filecategorystr)
            result_json= json.loads(status_json)
            print(iccid+":"+result_json['data']['msg'])
            if  result_json['status'] ==  "0" and result_json['code'] ==  "0":
               write_success_log(status_json)
               all_excel_log(iccid,custname,idcode,status_json)
            else:
               write_error_log(status_json)
               all_excel_log(iccid,custname,idcode,status_json)
               write_error_excel_log(iccid,custname,idcode,file1,file2,file3)
          else:
            status_json = error_json
            result_json= json.loads(status_json)
            print(iccid+":"+result_json['data']['msg'])
            write_error_log(status_json)
            all_excel_log(iccid,custname,idcode,status_json)
            write_error_excel_log(iccid,custname,idcode,file1,file2,file3)
       else:
          #第一次进来,重心赋值
          first_name = custname
          first_idcode = idcode
          print("正在进行:"+custname+"认证")
          #第一张
          basefile1 = get_basefile(file1)
          if basefile1 == '':
            error_json = '{"status": "0", "data": {"msg": "本地图片不存在"}, "code": "1"}'
            all_excel_log(iccid,custname,idcode,error_json)
            write_error_excel_log(iccid,custname,idcode,file1,file2,file3)
          else:
            md5basefile1 = get_base64file(basefile1)
            img_json1 = get_imgid(iccid,file1,basefile1,md5basefile1)
            result_data1 = json.loads(img_json1)
            #第二张
            basefile2 = get_basefile(file2)
            if basefile2 == '':
               error_json = '{"status": "0", "data": {"msg": "本地图片不存在"}, "code": "1"}'
               all_excel_log(iccid,custname,idcode,error_json)
               write_error_excel_log(iccid,custname,idcode,file1,file2,file3)
            else:
               md5basefile2 = get_base64file(basefile2)
               img_json2 = get_imgid(iccid,file2,basefile2,md5basefile2)
               result_data2 = json.loads(img_json2)
               #第三张
               basefile3 = get_basefile(file3)
               if basefile3 == '':
                  error_json = '{"status": "0", "data": {"msg": "本地图片不存在"}, "code": "1"}'
                  all_excel_log(iccid,custname,idcode,error_json)
                  write_error_excel_log(iccid,custname,idcode,file1,file2,file3)
               else:
                  md5basefile3 = get_base64file(basefile3)
                  img_json3 = get_imgid(iccid,file3,basefile3,md5basefile3)
                  result_data3 = json.loads(img_json3)
                  # 计算时间
                  if result_data1['status'] == "0" and result_data1['code'] == "0":
                     if  result_data2['status'] ==  "0" and result_data2['code'] ==  "0":
                           if  result_data3['status'] ==  "0" and result_data3['code'] ==  "0":
                              filecategorystr = result_data1['data']['msg']+','+result_data2['data']['msg']+','+result_data3['data']['msg']
                              status_json = get_status(iccid,idcode,custname,filecategorystr)
                              result_json = json.loads(status_json)
                              if  result_json['status'] ==  "0" and  result_json['code'] ==  "0":
                                 write_success_log(status_json)
                                 all_excel_log(iccid,custname,idcode,status_json)
                              else:
                                 write_error_log(status_json)
                                 all_excel_log(iccid,custname,idcode,status_json)
                                 write_error_excel_log(iccid,custname,idcode,file1,file2,file3)
                           else:
                              error_json = '{"status": "0", "data": {"msg": "'+result_data3['data']['msg']+'"}, "code": "1"}'
                              all_excel_log(iccid,custname,idcode,error_json)
                              write_error_excel_log(iccid,custname,idcode,file1,file2,file3)
                     else:
                           error_json = '{"status": "0", "data": {"msg": "'+result_data2['data']['msg']+'"}, "code": "1"}'
                           all_excel_log(iccid,custname,idcode,error_json)
                           write_error_excel_log(iccid,custname,idcode,file1,file2,file3)
                  else:
                     error_json = '{"status": "0", "data": {"msg": "'+result_data1['data']['msg']+'"}, "code": "1"}'
                     all_excel_log(iccid,custname,idcode,error_json)
                     write_error_excel_log(iccid,custname,idcode,file1,file2,file3)
print('end: ', tic())
